"""
parse_staff_targets.py
======================
Parses the 04_STAFF_TARGETS Excel template (the format with year section headers,
Office/Role/Partner columns, and — for zero targets).

Called by the /reference/staff-targets/upload endpoint.
Also used by import_engine_config.py as an alternative to reading the raw xlsm.

Format expected:
  Row 1: Title (ignored)
  Row 2: Instructions (ignored)
  Row 3: Spacer (ignored)
  Row 4: Column headers: Staff Member | Office | Role | Partner | Jan | Feb | ... | Dec
  Then repeating pattern:
    [Year row]     — merged cell with a 4-digit year
    [Data rows]    — one row per staff/office combo
"""

from typing import Dict, List, Tuple
import openpyxl


MONTHS = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]

# Canonical role → scheme mapping
ROLE_SCHEME_MAP = {
    "co_sub":  "CO_SUB",
    "co sub":  "CO_SUB",
    "sub":     "CO_SUB",
    "co":      "HCM_DIRECT",   # overridden to HN_DIRECT if office = HN or DN
    "direct":  "HCM_DIRECT",
}


def _s(v) -> str:
    if v is None: return ""
    return str(v).replace("\xa0"," ").strip()

def _i(v) -> int:
    s = _s(v).replace(",","").replace("—","0").replace("–","0").replace("-","0").strip()
    try:
        return max(0, int(float(s)))
    except:
        return 0

def _scheme(role: str, office: str) -> str:
    r = role.lower().strip()
    scheme = ROLE_SCHEME_MAP.get(r, "HCM_DIRECT")
    if scheme == "HCM_DIRECT" and office.upper() in ("HN", "DN"):
        scheme = "HN_DIRECT"
    return scheme


def parse_targets_excel(file_path: str) -> Tuple[List[dict], List[str]]:
    """
    Parse the staff targets template Excel file.

    Returns:
        records: list of dicts with keys:
            staff_name, office, role, scheme, partner,
            year, month (1-12), target (int)
        warnings: list of warning strings
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]

    records:  List[dict] = []
    warnings: List[str]  = []

    # ── Find header row ───────────────────────────────────────────────────────
    header_row = None
    col_map: Dict[str, int] = {}

    for i, row in enumerate(ws.iter_rows(max_row=8, values_only=True), start=1):
        hdrs = [_s(v).lower() for v in row]
        if "staff member" in hdrs:
            header_row = i
            for j, h in enumerate(hdrs):
                if "staff" in h:       col_map["staff"] = j
                elif "office" in h:    col_map["office"] = j
                elif "role" in h:      col_map["role"]   = j
                elif "partner" in h:   col_map["partner"] = j
                elif h in MONTHS:      col_map[h]        = j
            break

    if header_row is None:
        return [], ["Could not find header row. Expected row with 'Staff Member' column."]

    # ── Validate month columns ────────────────────────────────────────────────
    missing_months = [m for m in MONTHS if m not in col_map]
    if missing_months:
        warnings.append(f"Missing month columns: {', '.join(missing_months)}")

    # ── Parse data rows ───────────────────────────────────────────────────────
    current_year: int = 0
    data_start = header_row + 1

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(v for v in row): continue  # blank row

        col_a = _s(row[0])
        if not col_a: continue

        # Year section header — single 4-digit number in col A
        if col_a.isdigit() and len(col_a) == 4:
            current_year = int(col_a)
            continue

        # Skip legend / instructions rows
        if any(kw in col_a.lower() for kw in ("legend", "note", "upload", "instructions",
                                               "field", "rule", "scheme")):
            continue

        if current_year == 0:
            warnings.append(f"Skipping row (no year set yet): {col_a}")
            continue

        # ── Extract fields ────────────────────────────────────────────────────
        staff   = col_a.strip()
        office  = _s(row[col_map.get("office", -1)]).upper() if col_map.get("office") is not None and col_map.get("office","") != "" and col_map["office"] < len(row) else ""
        role    = _s(row[col_map.get("role",   -1)]) if col_map.get("role")    is not None and col_map.get("role","") != "" and col_map["role"] < len(row) else "CO"
        partner = _s(row[col_map.get("partner",-1)]) if col_map.get("partner") is not None and col_map.get("partner","") != "" and col_map.get("partner", 99) < len(row) else ""

        if not role: role = "CO"
        scheme = _scheme(role, office)

        # ── Month targets ─────────────────────────────────────────────────────
        has_data = False
        for m_idx, m_name in enumerate(MONTHS, start=1):
            col_idx = col_map.get(m_name)
            if col_idx is None or col_idx >= len(row): continue
            val = _i(row[col_idx])
            records.append({
                "staff_name": staff,
                "office":     office,
                "role":       role,
                "scheme":     scheme,
                "partner":    partner,
                "year":       current_year,
                "month":      m_idx,
                "target":     val,
            })
            if val > 0:
                has_data = True

        if not has_data:
            warnings.append(f"{current_year} | {staff} ({office or 'no office'}) — all targets are zero")

    return records, warnings


def aggregate_targets(records: List[dict]) -> Dict[Tuple[str,int,int], int]:
    """
    Returns summed targets per (canonical_staff_name, year, month).
    Staff with multiple office rows (e.g. Hoang Yen HCM + HN) are summed.
    The base name is normalised by stripping parenthetical office tags
    e.g. 'Trúc Quỳnh (HCM)' and 'Trúc Quỳnh (HN)' both → 'Trúc Quỳnh'.
    """
    import re
    totals: Dict[Tuple[str,int,int], int] = {}
    for r in records:
        name = re.sub(r'\s*\(.*?\)\s*', '', r["staff_name"]).strip()
        key  = (name, r["year"], r["month"])
        totals[key] = totals.get(key, 0) + r["target"]
    return totals


if __name__ == "__main__":
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else "staff_targets_template.xlsx"
    recs, warns = parse_targets_excel(path)
    print(f"Records: {len(recs)}")
    print(f"Warnings: {len(warns)}")
    for w in warns: print(f"  ! {w}")
    # Show aggregated totals sample
    totals = aggregate_targets(recs)
    print("\nSample aggregated targets (first 5):")
    for k, v in list(totals.items())[:5]:
        print(f"  {k} = {v}")
