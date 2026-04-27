# =============================================================================
# input.py  |  StudyLink Bonus Engine v1.0
# =============================================================================

from datetime import datetime, date
from typing import List, Tuple, Optional
import openpyxl

from .constants import *
from .config import BonusConfig
from .models import CaseRecord


def _s(v) -> str:
    if v is None: return ""
    return str(v).replace("\xa0", " ").strip()

def _d(v) -> Optional[date]:
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = _s(v).replace(" 00:00:00", "")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def _i(v) -> int:
    try: return int(float(str(v).replace(",","").replace(" ","").strip()))
    except: return 0


_HEADER_ALIASES = {
    "no.": "no", "student name": "student", "student id": "sid",
    "contract id": "contract", "contract signed date": "date",
    "contract date": "date", "client type": "ct",
    "country of study": "country", "country": "country",
    "refer source agent": "agent", "refer agent": "agent",
    "system type": "system",
    "application report status": "status",
    "application  report  status": "status",
    "application status": "status",                                # v7
    "visa received date": "visa_date", "visa date": "visa_date",
    "institution name": "institution",
    "course start date": "course_start", "course start": "course_start",
    "course status": "course_status",
    "counsellor name": "counsellor", "counsellor": "counsellor",   # v7
    "case officer name": "co", "case officer": "co",               # v7
    "notes": "notes",
    "bonus enrolled": "bonus",       # Manual report output
    "bonus  enrolled": "bonus",
    # ── V2 (30-column FIXED format) extension columns ────────────────────────
    # Present in V2 / v7 input files; absent from legacy V1 reports.
    # When missing, the existing classify.py inference pipeline runs unchanged.
    "pre-sales agent":          "presales_agent",
    "presales agent":           "presales_agent",
    "customer incentive (vnd)": "incentive",
    "customer incentive":       "incentive",
    "incentive (vnd)":          "incentive",
    "incentive":                "incentive",                       # v7
    "service fee type":         "service_fee_type",
    "deferral / waiver":        "deferral",
    "deferral/waiver":          "deferral",
    "deferral":                 "deferral",
    "package type":             "package_type",
    "office override":          "office_override",
    "handover":                 "handover",
    "target owner":             "target_owner",
    "case transition":          "case_transition",
    "prior month rate (vnd)":   "prior_month_rate",
    "prior month rate":         "prior_month_rate",
    "prior month rate / mgmt override amt":      "prior_month_rate",   # v7
    "prior month rate or mgmt override amt":     "prior_month_rate",   # v7 alt
    "prior month rate or\nmgmt override amt":    "prior_month_rate",   # v7 raw newline form
    "institution type":         "institution_type",
    "group/master agent name":  "group_agent_name",
    "group / master agent name":"group_agent_name",
    "group/master agent":       "group_agent_name",
    "group/agent name":         "group_agent_name",                # v7
    "group / agent name":       "group_agent_name",                # v7
    "group/agent":              "group_agent_name",                # v7
    "targets sheet name":       "targets_name",
    "targets sheet":            "targets_name",
    "targets name":             "targets_name",                    # v7
    # ── v7 ADDON-row fields (col 31-33) ──────────────────────────────────────
    "row type":                 "row_type",
    "row type (base/addon)":    "row_type",
    "row type\n(base/addon)":   "row_type",
    "add-on service code":      "addon_code",
    "add-on service\ncode":     "addon_code",
    "addon service code":       "addon_code",
    "add-on code":              "addon_code",
    "add-on count":             "addon_count",
    "add-on\ncount":            "addon_count",
    "addon count":              "addon_count",
    # ── v7 priority factor override (col 34, NEW Apr 2026) ────────────────────
    "priority factor":          "priority_factor",
    "priority factor override": "priority_factor",
    "priority factor\noverride":"priority_factor",
    "pri factor":               "priority_factor",
}


def _normalise_header(v) -> str:
    """
    Convert raw header cell value to a clean lowercase key for alias lookup.
    v7 template uses cells like '4\\nContract ID\\n[M]' or '17\\nPre-sales Agent\\nNEW'.
    Steps:
      1. Lowercase, replace non-breaking spaces.
      2. Remove a leading column number with optional newline (e.g. '4\\n').
      3. Strip trailing [M], [C], NEW markers.
      4. Collapse all whitespace (including newlines) to single spaces, trim.
    """
    import re
    s = _s(v).lower().replace("\xa0", " ")
    if not s:
        return s
    # Strip leading "4\n" or "27\n" etc.
    s = re.sub(r"^\s*\d+\s*[\n\r]+", "", s)
    # Strip trailing markers like [M], [C], NEW
    s = re.sub(r"\s*\[(m|c|m\?|c\?)\]\s*$", "", s)
    s = re.sub(r"\s*\bnew\b\s*$", "", s)
    # Collapse all whitespace to single spaces
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _detect_header(ws) -> Tuple[Optional[int], dict]:
    for i, row in enumerate(ws.iter_rows(max_row=6, values_only=True)):
        hdrs = [_normalise_header(v) for v in row]
        if "contract id" in hdrs or "no." in hdrs:
            col_map = {}
            for j, h in enumerate(hdrs):
                canon = _HEADER_ALIASES.get(h)
                if canon and canon not in col_map:
                    col_map[canon] = j
            return i, col_map
    return None, {}


def _get(row, col_map, key) -> str:
    idx = col_map.get(key)
    if idx is None or idx >= len(row): return ""
    return _s(row[idx])

def _get_date(row, col_map, key) -> Optional[date]:
    idx = col_map.get(key)
    if idx is None or idx >= len(row): return None
    return _d(row[idx])


def _v2_str(row, col_map, key, default: str = "") -> str:
    """Read a V2 column. Treats blank/'NONE' as 'not set' and returns default."""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return default
    val = _s(row[idx])
    if not val or val.upper() in ("NONE", ""):
        return default
    return val

def _v2_int(row, col_map, key, default: int = 0) -> int:
    """Read a V2 numeric column."""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return default
    return _i(row[idx])


def _v2_float(row, col_map, key, default: float = 0.0) -> float:
    """Read a V2 floating-point column (e.g., priority_factor 0.5/1.0)."""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    if val is None or val == "":
        return default
    try:
        if isinstance(val, (int, float)):
            return float(val)
        return float(str(val).strip().replace(",", "."))
    except (ValueError, TypeError):
        return default


def infer_institution_type(institution: str, system_type: str,
                           country: str = "") -> str:
    inst = institution or ""; sys = (system_type or "").lower()
    ctry = (country or "").lower()
    if "vietnam" in ctry or "viet nam" in ctry:
        if "rmit" in inst.lower(): return INST_RMIT_VN
        if "buv" in inst.lower() or "british university" in inst.lower(): return INST_BUV_VN
        return INST_OTHER_VN
    if "**" in inst: return INST_GROUP
    if "*"  in inst: return INST_MASTER_AGENT
    if "ngoài" in sys or "ngoai" in sys: return INST_OUT_OF_SYS
    return INST_DIRECT


def _dedup(cases: List[CaseRecord], cfg: BonusConfig) -> List[CaseRecord]:
    seen: dict = {}
    for i, c in enumerate(cases):
        if c.row_type == ROW_ADDON or not c.contract_id: continue
        cid = c.contract_id
        if cid not in seen:
            seen[cid] = i
        else:
            j = seen[cid]
            ri = cfg.get_status_rule(cases[j].app_status).dedup_rank
            rc = cfg.get_status_rule(c.app_status).dedup_rank
            if rc > ri:
                cases[j].is_duplicate = True; seen[cid] = i
            else:
                c.is_duplicate = True
    return cases


def parse_crm_report(file_path: str, cfg: BonusConfig
                     ) -> Tuple[List[CaseRecord], List[str]]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]
    hdr_idx, col_map = _detect_header(ws)
    if hdr_idx is None:
        return [], ["Could not detect header row"]

    cases: List[CaseRecord] = []
    warnings: List[str] = []

    for row in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
        no_val = _get(row, col_map, "no")
        if not no_val or cfg.is_skip_label(no_val): continue
        if not no_val.replace(".", "").isdigit(): continue

        contract    = _get(row, col_map, "contract")
        institution = _get(row, col_map, "institution")
        system      = _get(row, col_map, "system")
        country     = _get(row, col_map, "country")
        ct_text     = _get(row, col_map, "ct")

        c = CaseRecord(
            original_no   = no_val,
            student_name  = _get(row, col_map, "student"),
            student_id    = _get(row, col_map, "sid"),
            contract_id   = contract,
            contract_date = _get_date(row, col_map, "date"),
            client_type   = ct_text,
            country       = country,
            agent         = _get(row, col_map, "agent"),
            system_type   = system,
            app_status    = _get(row, col_map, "status"),
            visa_date     = _get_date(row, col_map, "visa_date"),
            institution   = institution,
            course_start  = _get_date(row, col_map, "course_start"),
            course_status = _get(row, col_map, "course_status"),
            counsellor    = _get(row, col_map, "counsellor"),
            case_officer  = _get(row, col_map, "co"),
            notes         = _get(row, col_map, "notes"),
            row_type      = ROW_BASE,
        )
        c.institution_type = infer_institution_type(institution, system, country)
        cr = cfg.get_country(country)
        c.country_code    = cr.code
        c.is_flat_country  = cr.is_flat_country
        c.is_vietnam       = cr.is_vietnam
        c.client_type_code = cfg.get_client_type_code(ct_text)
        # Detect agent-referred: external (non-StudyLink) agent in Refer
        # Source Agent field gives KPI weight 0.7 per 06_CLIENT_WEIGHTS
        # sub-referral rule, AND blocks priority bonus on direct schemes.
        # Internal-agent patterns are loaded from ref_internal_agents
        # (see cfg.internal_agent_patterns). The team can add new
        # internal office patterns there without code changes.
        agent_lower = c.agent.lower()
        c.is_agent_referred = (bool(c.agent) and len(c.agent) > 2 and
                               not any(x in agent_lower
                                       for x in cfg.internal_agent_patterns))

        # ── V2 (30-column FIXED format) field population ─────────────────────
        # When V2 columns are present in the file, their explicit values
        # override the engine's inferred values. When V2 columns are absent
        # (legacy V1 17-column file), these calls are no-ops and the existing
        # classify.py inference pipeline runs unchanged downstream.
        c.presales_agent  = _v2_str(row, col_map, "presales_agent", c.presales_agent)
        c.incentive       = _v2_int(row, col_map, "incentive", c.incentive)
        c.service_fee_type = _v2_str(row, col_map, "service_fee_type", c.service_fee_type)
        c.deferral        = _v2_str(row, col_map, "deferral", c.deferral)
        c.package_type    = _v2_str(row, col_map, "package_type", c.package_type)
        office_override   = _v2_str(row, col_map, "office_override", "")
        if office_override:
            c.office = office_override.upper()
        c.handover        = _v2_str(row, col_map, "handover", c.handover)
        c.target_owner    = _v2_str(row, col_map, "target_owner", c.target_owner)
        c.case_transition = _v2_str(row, col_map, "case_transition", c.case_transition)
        c.prior_month_rate = _v2_int(row, col_map, "prior_month_rate", c.prior_month_rate)
        inst_type_v2      = _v2_str(row, col_map, "institution_type", "")
        if inst_type_v2:
            c.institution_type = inst_type_v2  # V2 explicit value beats inferred
        c.group_agent_name = _v2_str(row, col_map, "group_agent_name", c.group_agent_name)
        c.targets_name    = _v2_str(row, col_map, "targets_name", c.targets_name)
        c.priority_factor = _v2_float(row, col_map, "priority_factor", c.priority_factor)

        cases.append(c)

    cases = _dedup(cases, cfg)
    n = 0
    for c in cases:
        if not c.is_duplicate:
            n += 1; c.display_no = n
    for c in cases:
        warnings.extend(c.warn_flags)
    return cases, warnings


def read_manual_report(file_path: str) -> Tuple[int, dict]:
    """
    Reads manual bonus report. Returns (total, {contract_id: bonus}).
    Handles different header layouts across files.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]
    hdr_idx, col_map = _detect_header(ws)
    if hdr_idx is None:
        return 0, {}

    contract_col = col_map.get("contract")
    bonus_col    = col_map.get("bonus")
    if contract_col is None or bonus_col is None:
        return 0, {}

    total = 0
    by_contract = {}
    for row in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
        if contract_col >= len(row): continue
        contract = _s(row[contract_col])
        if not contract: continue
        if contract.lower() in ("tổng", "tong", "total"): break
        # Try bonus_col and bonus_col+1 to handle column-offset layouts
        b = 0
        for try_col in [bonus_col, bonus_col + 1]:
            if try_col < len(row):
                try:
                    v = int(float(_s(row[try_col]).replace(",","")))
                    if v != 0:
                        b = v
                        break
                except: pass
        if contract.startswith("SLC-"):
            by_contract[contract] = b
            total += b
    return total, by_contract
