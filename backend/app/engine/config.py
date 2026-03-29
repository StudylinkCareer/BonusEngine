# =============================================================================
# config.py  |  StudyLink Bonus Engine v1.0
# Replaces: modConfig.bas
# =============================================================================

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
import unicodedata
import openpyxl

from .constants import *


@dataclass
class StatusRule:
    status: str
    counts_as_enrolled: bool = False
    coun_pct: float = 0.0
    co_direct_pct: float = 0.0
    co_sub_pct: float = 0.0
    is_carry_over: bool = False
    is_current_enrolled: bool = False
    is_zero_bonus: bool = False
    fees_paid_non_enrolled: bool = False
    is_visa_granted: bool = False
    dedup_rank: int = 0


@dataclass
class StaffTarget:
    name: str
    office: str = OFFICE_HCM
    role: str = "CO"
    scheme: str = SCHEME_HCM_DIRECT
    targets: Dict[int, Dict[int, int]] = field(default_factory=dict)


@dataclass
class ServiceFeeRule:
    code: str
    coun_bonus: int = 0
    co_bonus: int = 0
    active: bool = True
    category: str = ""
    applies_to: str = "ALL"
    description: str = ""


@dataclass
class CountryRule:
    crm_text: str
    code: str
    is_flat_country: bool = False
    is_vietnam: bool = False


@dataclass
class PriorityInstitution:
    name: str
    bonus_pct: float = 0.0
    annual_target: int = 0
    achieved_ytd: int = 0


class BonusConfig:
    def __init__(self):
        self.status_rules:    Dict[str, StatusRule]      = {}
        self.staff_targets:   Dict[str, StaffTarget]     = {}
        self.service_fees:    Dict[str, ServiceFeeRule]  = {}
        self.country_codes:   Dict[str, CountryRule]     = {}
        self.client_types:    Dict[str, str]             = {}
        self.staff_name_map:  Dict[str, str]             = {}
        self.skip_labels:     frozenset                  = SKIP_LABELS
        self.master_agents:   List[str]                  = []
        self.priority_instns: List[PriorityInstitution]  = []
        self.base_rates:      Dict[str, Dict[str, int]]  = {}

    def get_status_rule(self, status: str) -> StatusRule:
        return self.status_rules.get(status.strip().lower(),
               StatusRule(status=status, is_zero_bonus=True))

    def get_service_fee(self, code: str, category: str = "") -> Optional[ServiceFeeRule]:
        if not code or code.upper() in (SVC_NONE, ""):
            return None
        r = self.service_fees.get(code.strip().lower())
        if r is None:
            return None
        if category and r.category.upper() != category.upper():
            return None
        if not r.active:
            return None
        return r

    def get_service_fee_any(self, code: str) -> Optional[ServiceFeeRule]:
        return self.service_fees.get(code.strip().lower()) if code else None

    def resolve_staff_name(self, crm_name: str) -> str:
        return self.staff_name_map.get(crm_name.strip().lower(), crm_name.strip())

    @staticmethod
    def _ascii(s: str) -> str:
        """Strips diacritics for fuzzy name matching."""
        return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii').lower().strip()

    def get_staff_target(self, name: str, year: int, month: int) -> Tuple[int, str]:
        """Returns (target, scheme). Tries exact, resolved, ascii-normalised, then partial."""
        resolved = self.resolve_staff_name(name)
        # Exact match attempts (fast path)
        for key in [resolved.lower(), name.strip().lower()]:
            st = self.staff_targets.get(key)
            if st is not None:
                return st.targets.get(year, {}).get(month, 0), st.scheme
        # ASCII-normalised match (strips diacritics — "Gia Mẫn" == "Gia Man")
        name_ascii     = self._ascii(name)
        resolved_ascii = self._ascii(resolved)
        for key, st in self.staff_targets.items():
            if self._ascii(key) in (name_ascii, resolved_ascii):
                return st.targets.get(year, {}).get(month, 0), st.scheme
        # Substring partial match (last resort)
        for key, st in self.staff_targets.items():
            ka = self._ascii(key)
            if ka in name_ascii or name_ascii in ka:
                return st.targets.get(year, {}).get(month, 0), st.scheme
        return 0, SCHEME_HCM_DIRECT

    def get_staff_scheme(self, name: str) -> str:
        resolved = self.resolve_staff_name(name)
        for key in [resolved.lower(), name.strip().lower()]:
            st = self.staff_targets.get(key)
            if st is not None:
                return st.scheme
        return SCHEME_HCM_DIRECT

    def get_base_rate(self, scheme: str, tier: str) -> int:
        return self.base_rates.get(scheme, {}).get(tier, 0)

    def get_base_rate_coun(self, scheme: str, tier: str) -> int:
        return self.base_rates.get(scheme, {}).get(f"coun_{tier}", 0)

    def get_country(self, crm_text: str) -> CountryRule:
        return self.country_codes.get(crm_text.strip().lower(),
               CountryRule(crm_text=crm_text, code=crm_text))

    def get_client_type_code(self, crm_text: str) -> str:
        return self.client_types.get(crm_text.strip().lower(), "")

    def is_skip_label(self, label: str) -> bool:
        return label.strip().lower() in self.skip_labels

    def get_kpi_weight(self, ct_code: str, inst_type: str, scheme: str) -> float:
        if ct_code in (CT_SUMMER, CT_GUARDIAN, CT_TOURIST,
                       CT_MIGRATION, CT_DEPENDANT, CT_VISA_ONLY):
            return 0.0
        if ct_code == CT_VIETNAM:
            return 0.5
        if inst_type == INST_MASTER_AGENT:
            return 1.0
        if inst_type == INST_OUT_OF_SYS:
            return 0.7
        # CO_SUB: each enrolled case counts as 1.0 — the sub-referral 0.7 weight
        # applies to CO_DIRECT cases referred via sub-agent, not to the CO_SUB role itself
        return 1.0


def load_config(workbook_path: str) -> BonusConfig:
    cfg = BonusConfig()
    wb  = openpyxl.load_workbook(workbook_path, data_only=True)
    _load_status_rules(wb, cfg)
    _load_base_rates(wb, cfg)
    _load_service_fees(wb, cfg)
    _load_staff_targets(wb, cfg)
    _load_staff_names(wb, cfg)
    _load_countries(wb, cfg)
    _load_client_types(wb, cfg)
    _load_priority_instns(wb, cfg)
    return cfg


def _s(v) -> str:
    return str(v).strip() if v is not None else ""

def _i(v) -> int:
    try:
        return int(float(str(v).replace("—","0").replace("–","0").replace(",","").strip()))
    except (ValueError, TypeError):
        return 0

def _f(v) -> float:
    s = str(v or "0").replace("%","").strip()
    try:
        f = float(s)
        return f / 100 if f > 1 else f
    except (ValueError, TypeError):
        return 0.0

def _b(v) -> bool:
    return str(v or "").strip().upper() in ("Y", "YES", "TRUE", "1")


def _load_status_rules(wb, cfg):
    ws = wb[WS_STATUS_RULES]
    for row in ws.iter_rows(min_row=3, values_only=True):
        status = _s(row[0])
        if not status or status.startswith("("):
            continue
        cfg.status_rules[status.lower()] = StatusRule(
            status=status,
            counts_as_enrolled   =_b(row[2]),
            coun_pct             =_f(row[3]),
            co_direct_pct        =_f(row[4]),
            co_sub_pct           =_f(row[5]),
            is_carry_over        =_b(row[6]),
            is_current_enrolled  =_b(row[7]),
            is_zero_bonus        =_b(row[8]),
            fees_paid_non_enrolled=_b(row[9]),
            is_visa_granted      =_b(row[10]),
            dedup_rank           =_i(row[11]) if len(row) > 11 else 0,
        )


def _load_base_rates(wb, cfg):
    ws   = wb[WS_BASE_RATES]
    rows = list(ws.iter_rows(values_only=True))

    def r(idx, col, default=0):
        try: return _i(rows[idx][col]) or default
        except: return default

    # HCM Direct (row 5 = index 4)
    cfg.base_rates[SCHEME_HCM_DIRECT] = {
        "out_sys_co":            r(4,2),
        "visa_only_co":          r(4,3),
        f"coun_{TIER_UNDER}":    r(4,4),
        TIER_UNDER:              r(4,5),
        f"coun_{TIER_MEET_HIGH}":r(4,6),
        TIER_MEET_HIGH:          r(4,7),
        f"coun_{TIER_MEET_LOW}": r(4,8),
        TIER_MEET_LOW:           r(4,9),
        f"coun_{TIER_OVER}":     r(4,10),
        TIER_OVER:               r(4,11),
    }
    # HN/DN Direct (row 37 = index 36)
    cfg.base_rates[SCHEME_HN_DIRECT] = {
        "out_sys_co":            r(36,2),
        "visa_only_co":          r(36,3),
        f"coun_{TIER_UNDER}":    r(36,4),
        TIER_UNDER:              r(36,5),
        f"coun_{TIER_MEET_HIGH}":r(36,6),
        TIER_MEET_HIGH:          r(36,7),
        f"coun_{TIER_MEET_LOW}": r(36,8),
        TIER_MEET_LOW:           r(36,9),
        f"coun_{TIER_OVER}":     r(36,10),
        TIER_OVER:               r(36,11),
    }
    # CO Sub (row 22 = index 21)
    # Section B column layout: Fixed | EnrolUnder | EnrolMeet | EnrolOver | FullUnder | FullMeet | FullOver
    # Truong An uses "Enrolment only in-system" → cols 2,3,4
    cfg.base_rates[SCHEME_CO_SUB] = {
        "fixed":        r(21,1),   # 400k out-sys fixed
        TIER_UNDER:     r(21,2),   # 700k enrol-only under
        TIER_MEET_HIGH: r(21,3),   # 900k enrol-only meet
        TIER_MEET_LOW:  r(21,3),   # 900k enrol-only meet
        TIER_OVER:      r(21,4),   # 1,100k enrol-only over
        # Vietnam special rates for CO_SUB (row 24 index 23, col 2)
        "rmit_vn":      r(23,2),   # 600k
        "other_vn":     r(24,2),   # 300k (Other VN)
        "summer":       r(25,2),   # 300k
        "flat":         r(22,1),   # 300k (Thai/Phil/ML fixed)
    }


def _load_service_fees(wb, cfg):
    ws = wb[WS_SERVICE_FEES]
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = _s(row[0])
        if not code or code.startswith("("):
            continue
        cat = _s(row[5]).upper()
        if cat not in (SVC_SERVICE_FEE, SVC_PACKAGE, SVC_CONTRACT, SVC_ADDON):
            continue
        cfg.service_fees[code.lower()] = ServiceFeeRule(
            code=code, coun_bonus=_i(row[2]), co_bonus=_i(row[3]),
            active=_b(row[4]) if len(row) > 4 else True,
            category=cat,
            applies_to=_s(row[6]).upper() if len(row) > 6 else "ALL",
            description=_s(row[8]) if len(row) > 8 else "",
        )
    # Inject VISA_RENEWAL_SUPPORT if not in workbook yet
    if "visa_renewal_support" not in cfg.service_fees:
        cfg.service_fees["visa_renewal_support"] = ServiceFeeRule(
            code="VISA_RENEWAL_SUPPORT", coun_bonus=0, co_bonus=250000,
            active=True, category=SVC_SERVICE_FEE, applies_to="ALL",
            description="Student visa renewal support AUS/NZ (50% when handover)"
        )
    # Inject superior Package 8tr if not present
    if "superior package 8tr" not in cfg.service_fees:
        cfg.service_fees["superior package 8tr"] = ServiceFeeRule(
            code="superior Package 8tr", coun_bonus=1000000, co_bonus=500000,
            active=True, category=SVC_PACKAGE, applies_to="DIRECT",
            description="Superior Package 8M (Singapore/Germany)"
        )


def _load_staff_targets(wb, cfg):
    ws = wb[WS_TARGETS]
    current_year = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        a = _s(row[0]).strip("'")
        if not a:
            continue
        if a.isdigit():
            current_year = int(a)
            continue
        if current_year is None:
            continue
        # Strip None values from merged cells
        clean = [v for v in row if v is not None]
        if len(clean) < 2:
            continue
        staff = _s(clean[0]).strip("'")
        b = _s(clean[1]).strip("'").upper()
        if b in (OFFICE_HCM, OFFICE_HN, OFFICE_DN):
            office = b
            role = _s(clean[2]).strip("'").upper() if len(clean) > 2 else "CO"
            # If index 3 is numeric (or '—'), no Partner column → months start at 3 (2024 style)
            # If index 3 is a name string → Partner column present → months start at 4 (2025 style)
            idx3 = _s(clean[3]).strip("'") if len(clean) > 3 else "0"
            try:
                float(idx3.replace("—", "0").replace("–", "0"))
                months = clean[3:]      # 2024 style: Staff | Office | Role | Jan...
            except (ValueError, TypeError):
                months = clean[4:]      # 2025 style: Staff | Office | Role | Partner | Jan...
        elif b in ("CO_SUB", "DIRECT"):
            office = OFFICE_HCM
            role = b
            months = clean[2:]
        else:
            office = OFFICE_HCM
            role = b
            months = clean[3:]
        scheme = (SCHEME_CO_SUB if role == "CO_SUB"
                  else SCHEME_HN_DIRECT if office in (OFFICE_HN, OFFICE_DN)
                  else SCHEME_HCM_DIRECT)
        key = staff.lower()
        if key not in cfg.staff_targets:
            cfg.staff_targets[key] = StaffTarget(
                name=staff, office=office, role=role, scheme=scheme)
        st = cfg.staff_targets[key]
        if current_year not in st.targets:
            st.targets[current_year] = {}
        for m in range(1, 13):
            val = _i(months[m - 1]) if m - 1 < len(months) else 0
            st.targets[current_year][m] = st.targets[current_year].get(m, 0) + val


def _load_staff_names(wb, cfg):
    ws = wb[WS_STAFF_NAMES]
    for row in ws.iter_rows(min_row=5, values_only=True):
        crm = _s(row[1]) if len(row) > 1 else ""
        tgt = _s(row[2]) if len(row) > 2 else ""
        if crm and tgt:
            cfg.staff_name_map[crm.lower()] = tgt


def _load_countries(wb, cfg):
    ws = wb[WS_COUNTRIES]
    for row in ws.iter_rows(min_row=4, values_only=True):
        crm = _s(row[0])
        if not crm:
            continue
        cfg.country_codes[crm.lower()] = CountryRule(
            crm_text=crm, code=_s(row[1]),
            is_flat_country=_b(row[2]),
            is_vietnam=_b(row[3]) if len(row) > 3 else False,
        )


def _load_client_types(wb, cfg):
    ws = wb[WS_CLIENT_TYPES]
    for row in ws.iter_rows(min_row=4, values_only=True):
        crm = _s(row[0]); code = _s(row[1])
        if crm and code:
            cfg.client_types[crm.lower()] = code


def _load_priority_instns(wb, cfg):
    if WS_PRIORITY not in wb.sheetnames:
        return
    ws = wb[WS_PRIORITY]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _s(row[0])
        if not name or name.startswith("("):
            continue
        cfg.priority_instns.append(PriorityInstitution(
            name=name, bonus_pct=_f(row[1]),
            annual_target=_i(row[2]) if len(row) > 2 else 0,
        ))
