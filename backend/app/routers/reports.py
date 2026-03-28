# =============================================================================
# routers/reports.py
# Run history, results display, and Excel export endpoints.
# =============================================================================

import io
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Run, Case, User
from ..schemas import RunOut, RunSummary
from ..routers.auth import get_current_user

router = APIRouter()


@router.get("/", response_model=List[RunSummary])
def list_runs(
    staff_name: Optional[str] = Query(None),
    run_year: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns run history. Admins see all runs.
    Non-admins see only their own staff's runs.
    """
    query = db.query(Run)

    # Non-admins restricted to their own staff name
    if not current_user.is_admin and current_user.staff_name:
        query = query.filter(Run.staff_name == current_user.staff_name)
    elif staff_name:
        query = query.filter(Run.staff_name == staff_name)

    if run_year:
        query = query.filter(Run.run_year == run_year)
    if status:
        query = query.filter(Run.status == status)

    return query.order_by(Run.run_year.desc(), Run.run_month.desc()).all()


@router.get("/{run_id}", response_model=RunOut)
def get_run(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Returns full run detail including all cases and their bonus results."""
    run = db.query(Run).filter(Run.id == run_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    return run


@router.get("/{run_id}/export")
def export_run(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Exports the run results as a formatted Excel file.
    Matches the layout of the VBA engine output tab.
    """
    run = db.query(Run).filter(Run.id == run_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    if run.status not in ("calculated", "approved"):
        raise HTTPException(
            status_code=400,
            detail="Run must be calculated before export"
        )

    cases = db.query(Case).filter(Case.run_id == run_id).all()

    excel_bytes = _build_excel(run, cases)

    filename = (f"BonusReport_{run.staff_name.replace(' ', '_')}_"
                f"{run.run_month:02d}{run.run_year}.xlsx")

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_excel(run: Run, cases: list) -> bytes:
    """Builds a formatted Excel output matching the VBA engine output layout."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{run.staff_name[:15]}_{run.run_month:02d}{run.run_year}"

    # Colours
    DARK_BLUE  = "1E4E79"
    MID_BLUE   = "2E75B6"
    LIGHT_BLUE = "BDD7EE"
    AMBER      = "FFFF8C"
    GREEN      = "E2EFDA"
    WHITE      = "FFFFFF"

    def style_cell(cell, bold=False, font_color=WHITE, bg=None, size=9, wrap=False, align="left"):
        cell.font = Font(name="Arial", size=size, bold=bold, color=font_color)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)

    # Row 1: Title
    ws.merge_cells(f"A1:AP1")
    ws["A1"] = (f"Bao cao Ho so - {run.staff_name} - "
                f"Thang {run.run_month:02d}/{run.run_year}")
    style_cell(ws["A1"], bold=True, bg=DARK_BLUE, size=11)
    ws.row_dimensions[1].height = 24

    # Row 2: Subtitle
    ws.merge_cells("A2:AP2")
    ws["A2"] = (f"Staff: {run.staff_name}   |   "
                f"Period: {_month_name(run.run_month)} {run.run_year}")
    style_cell(ws["A2"], bg=LIGHT_BLUE, font_color="000000", size=9)

    # Row 3: Spacer
    ws.row_dimensions[3].height = 6

    # Row 4: Column headers
    headers = [
        "No.", "Student Name", "Student ID", "Contract ID", "Contract Date",
        "Client Type", "Country", "Agent", "System", "Status",
        "Visa Date", "Institution", "Course Start", "Course Status",
        "Counsellor", "CO", "Pre-sales", "Incentive", "Notes",
        "Service Fee", "Deferral", "Package", "Office", "Handover",
        "Target Owner", "Case Trans.", "Prior Rate", "Inst. Type",
        "Group Agent", "Targets Name", "Row Type", "Addon Code", "Addon Count",
        "BONUS Enrolled", "Note Enrolled", "Note2", "BONUS Priority",
        "Note Priority", "Note2", "Prior Advances", "Net Payable", "Base Rate",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        style_cell(cell, bold=True, bg=MID_BLUE, wrap=True)

    ws.row_dimensions[4].height = 30

    # Data rows
    row_num = 5
    total_enr = 0
    total_pri = 0

    for case in cases:
        values = [
            case.original_no, case.student_name, case.student_id,
            case.contract_id, case.contract_date, case.client_type,
            case.country, case.agent, case.system_type, case.app_status,
            case.visa_date, case.institution, case.course_start,
            case.course_status, case.counsellor, case.case_officer,
            case.presales_agent, case.incentive, case.notes,
            case.service_fee_type, case.deferral, case.package_type,
            case.office_override, case.handover, case.target_owner,
            case.case_transition, case.prior_month_rate, case.institution_type,
            case.group_agent_name, case.targets_name, case.row_type,
            case.addon_code, case.addon_count,
            case.bonus_enrolled, case.note_enrolled, case.note_enrolled2,
            case.bonus_priority, case.note_priority, case.note_priority2,
            case.prior_advances, case.net_payable, case.stored_base_rate,
        ]

        is_addon = (case.row_type or "").upper() == "ADDON"
        bg = AMBER if is_addon else None

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            if col_idx in (18, 27, 34, 37, 40, 41, 42) and isinstance(val, int):
                cell.number_format = "#,##0"

        if not is_addon:
            total_enr += case.bonus_enrolled or 0
            total_pri += case.bonus_priority or 0

        row_num += 1

    # Total row
    ws.cell(row=row_num, column=1, value="TONG").font = Font(name="Arial", bold=True, size=9)
    total_cell = ws.cell(row=row_num, column=34, value=total_enr + total_pri)
    total_cell.font = Font(name="Arial", bold=True, size=9)
    total_cell.number_format = "#,##0"

    # Column widths
    widths = {
        1: 4, 2: 22, 3: 12, 4: 13, 5: 12, 6: 20, 7: 12,
        8: 16, 9: 14, 10: 28, 11: 12, 12: 30, 13: 12, 14: 14,
        15: 20, 16: 20, 17: 18, 18: 14, 19: 28, 20: 18, 21: 18,
        22: 18, 23: 12, 24: 10, 25: 18, 26: 12, 27: 14, 28: 20,
        29: 20, 30: 16, 31: 12, 32: 20, 33: 10, 34: 18, 35: 30,
        36: 20, 37: 18, 38: 30, 39: 20, 40: 16, 41: 16, 42: 16,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _month_name(month: int) -> str:
    names = ["", "January", "February", "March", "April", "May", "June",
             "July", "August", "September", "October", "November", "December"]
    return names[month] if 1 <= month <= 12 else str(month)
