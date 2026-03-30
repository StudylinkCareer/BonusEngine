"""
import_engine_config.py
=======================
Imports ALL configuration from engine.xlsm into PostgreSQL.
Zero hardcoded rates — everything goes into the DB.
Safe to re-run — all tables are cleared and re-imported each time.

Run from backend/ folder:
    python import_engine_config.py path/to/engine.xlsm
"""

import sys, os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import json
import openpyxl
from datetime import date, datetime
from app.database import engine as db_engine, SessionLocal
from app.models import (
    Base, StaffName, StaffTarget, MasterAgent, CountryCode,
    ClientTypeMap, StatusRule, ServiceFeeRate, ReferenceList,
    PriorityInstitution, YtdTracker, ClientWeight, ContractBonus,
    AdvancePayment, BaseRate, IncentiveTier, SpecialRate,
    CountryRate, PartnerInstitution, AdvanceRule
)

EFFECTIVE_DATE = date(2024, 6, 1)  # Policy effective date from policy summary


def _s(v) -> str:
    if v is None: return ""
    return str(v).replace("\xa0"," ").replace("?"," ").strip()

def _i(v) -> int:
    try:
        return int(float(str(v).replace("—","0").replace("–","0").replace(",","").replace("--","0").strip()))
    except:
        return 0

def _f(v) -> float:
    s = str(v or "0").replace("%","").replace("--","0").strip()
    try:
        f = float(s)
        return f/100 if f>1 else f
    except:
        return 0.0

def _b(v) -> bool:
    return str(v or "").strip().upper() in ("Y","YES","TRUE","1")

def _dt(v):
    if v is None: return None
    if isinstance(v, datetime): return v
    try:
        s = str(v).strip().replace(" 00:00:00","")
        for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y"):
            try: return datetime.strptime(s, fmt)
            except: pass
    except: pass
    return None

def _is_rate(v) -> bool:
    """True if the cell has a numeric rate (not -- or None)."""
    if v is None: return False
    s = str(v).strip()
    return s not in ("","--","—","–","None") and any(c.isdigit() for c in s)


def import_all(xlsm_path: str):
    Base.metadata.create_all(bind=db_engine)
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(xlsm_path, data_only=True)
        print(f"Opened: {xlsm_path}\n")

        import_staff_names(wb, db)
        import_staff_targets(wb, db)
        import_country_codes(wb, db)
        import_client_type_map(wb, db)
        import_status_rules(wb, db)
        import_service_fee_rates(wb, db)
        import_master_agents(wb, db)
        import_skip_labels(wb, db)
        import_priority_institutions(wb, db)
        import_ytd_tracker(wb, db)
        import_client_weights(wb, db)
        import_contract_bonuses(wb, db)
        import_advance_tracker(wb, db)
        import_base_rates(wb, db)        # NEW
        import_special_rates(wb, db)     # NEW
        import_incentive_tiers(db)       # NEW — seeds defaults
        import_country_rates(wb, db)     # NEW
        import_partner_instns(db)        # NEW — seeds defaults
        import_advance_rules(db)         # NEW — seeds defaults

        db.commit()
        print("\n✅ All tables imported successfully!")
    except Exception as e:
        db.rollback()
        print(f"\n❌ Import failed: {e}")
        import traceback; traceback.print_exc()
        raise
    finally:
        db.close()


# =============================================================================
# 12_STAFF_NAMES
# =============================================================================
def import_staff_names(wb, db):
    ws = wb["12_STAFF_NAMES"]
    db.query(ReferenceList).filter(ReferenceList.list_name=="staff_name_map").delete()
    db.flush()
    count=0; seen_canonical=set(); seen_crm=set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        crm=_s(row[1]); canonical=_s(row[2])
        role_note=_s(row[4]).lower() if len(row)>4 else ""
        if not crm or not canonical or crm in seen_crm: continue
        seen_crm.add(crm)
        db.add(ReferenceList(list_name="staff_name_map", value=crm, sort_order=count, is_active=True))
        if canonical not in seen_canonical:
            seen_canonical.add(canonical)
            if not db.query(StaffName).filter(StaffName.full_name==canonical).first():
                office="HCM"
                if "hn" in role_note or "ha noi" in role_note: office="HN"
                elif "dn" in role_note or "da nang" in role_note: office="DN"
                role="counsellor"
                if "co_sub" in role_note: role="case_officer"
                elif "presales" in role_note: role="presales"
                scheme="CO_SUB" if role=="case_officer" else \
                       "HN_DIRECT" if office in ("HN","DN") else "HCM_DIRECT"
                db.add(StaffName(full_name=canonical, short_name=canonical, office=office,
                                 role=role, scheme=scheme, is_active=True,
                                 start_date=EFFECTIVE_DATE))
        count+=1
    db.flush()
    print(f"  12_STAFF_NAMES:       {count} CRM mappings, {len(seen_canonical)} canonical names")


# =============================================================================
# 04_STAFF_TARGETS
# =============================================================================
def import_staff_targets(wb, db):
    ws = wb["04_STAFF_TARGETS"]
    db.query(StaffTarget).delete(); db.flush()
    current_year=None; count=0
    for row in ws.iter_rows(min_row=3, values_only=True):
        a=_s(row[0]).strip("'")
        if not a: continue
        if a.isdigit() and len(a)==4: current_year=int(a); continue
        if current_year is None or a.lower() in ("staff member","name"): continue
        idx3=_s(row[3]) if len(row)>3 and row[3] is not None else "0"
        try:
            float(idx3.replace("—","0").replace("–","0"))
            months=list(row[3:15])
        except:
            months=list(row[4:16])
        office=_s(row[1]) or "HCM"
        if office not in ("HCM","HN","DN"): office="HCM"
        for m in range(1,13):
            val=_i(months[m-1]) if m-1<len(months) else 0
            if val>0:
                db.add(StaffTarget(staff_name=a, office=office, month=m, year=current_year, target=val))
                count+=1
    db.flush()
    print(f"  04_STAFF_TARGETS:     {count} monthly target records")


# =============================================================================
# 14_COUNTRY_CODES
# =============================================================================
def import_country_codes(wb, db):
    ws = wb["14_COUNTRY_CODES"]
    db.query(CountryCode).delete(); db.flush()
    count=0; seen=set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        crm=_s(row[0])
        if not crm or crm.startswith("(") or crm in seen: continue
        seen.add(crm); code=_s(row[1])
        region="Asia"
        if code in ("AU","NZ"): region="Oceania"
        elif code in ("CA","US","GD"): region="Americas"
        elif code in ("GB","IE","FR","DE","NL","SE","NO","DK","FI","CH","CZ","HU","PL","AT","BE","ES","IT"): region="Europe"
        db.add(CountryCode(country_name=crm, country_code=code, region=region, is_active=True))
        count+=1
    db.flush()
    print(f"  14_COUNTRY_CODES:     {count} records")


# =============================================================================
# 15_CLIENT_TYPE_MAP
# =============================================================================
def import_client_type_map(wb, db):
    ws = wb["15_CLIENT_TYPE_MAP"]
    db.query(ClientTypeMap).delete(); db.flush()
    count=0; seen=set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        raw=_s(row[0]); canonical=_s(row[1])
        if not raw or not canonical or raw.startswith("(") or raw in seen: continue
        seen.add(raw)
        db.add(ClientTypeMap(raw_value=raw, canonical=canonical, display_name=raw, is_active=True))
        count+=1
    db.flush()
    print(f"  15_CLIENT_TYPE_MAP:   {count} records")


# =============================================================================
# 05_STATUS_RULES — now includes split percentages
# =============================================================================
def import_status_rules(wb, db):
    ws = wb["05_STATUS_RULES"]
    db.query(StatusRule).delete()
    db.query(ReferenceList).filter(ReferenceList.list_name=="application_status").delete()
    db.flush()

    status_order = [
        "Closed - Visa granted, then enrolled",
        "Closed - Visa granted (visa only)",
        "Closed - Visa granted then cancelled",
        "Closed - Visa refused",
        "Closed - Visa refused then granted",
        "Closed - Enrolment (only)",
        "Closed - Enrolled then cancelled",
        "Closed - Enrolled, then Visa granted",
        "Closed - Cancelled",
        "Current - Enrolled",
        "Current - Visa refused",
        "Pending - Visa refused",
        "Closed - Institution refused",
        "Closed - Visa granted",
        "Closed - Enrolled then visa refused",
        "Closed - Enrolment",
    ]

    # Dedup rank lookup
    rank_map = {
        "closed - visa granted, then enrolled": 5,
        "closed - enrolment": 5,
        "closed - enrolment (only)": 5,
        "closed - visa refused then granted": 5,
        "closed - enrolled, then visa granted": 4,
        "current - enrolled": 2,
        "closed - cancelled": 1,
        "closed - visa refused": 1,
        "closed - enrolled then cancelled": 1,
        "closed - institution refused": 1,
        "closed - enrolled then visa refused": 1,
        "current - visa refused": 1,
        "pending - visa refused": 1,
    }

    # col layout: 0=status, 1=activities, 2=counts_enrolled, 3=coun%, 4=co_direct%, 5=co_sub%,
    #             6=is_carryover, 7=is_current_enrolled, 8=is_zero_bonus, 9=fees_paid_non_enrolled
    count=0; seen=set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        status=_s(row[0])
        if not status or status in seen: continue
        if not any(status.startswith(p) for p in ("Closed","Current","Pending","Bundled")): continue
        seen.add(status)

        counts_enrol_raw = _s(row[2]).upper() if len(row)>2 else ""
        counts_enrol = counts_enrol_raw in ("YES","YES (50%)")
        coun_pct   = _f(row[3]) if len(row)>3 and _is_rate(row[3]) else 0.0
        co_dir_pct = _f(row[4]) if len(row)>4 and _is_rate(row[4]) else 0.0
        co_sub_pct = _f(row[5]) if len(row)>5 and _is_rate(row[5]) else 0.0
        is_co      = _b(row[6]) if len(row)>6 else False  # IsCarryOver
        is_cur     = _b(row[7]) if len(row)>7 else False  # IsCurrentEnrolled
        is_zero    = _b(row[8]) if len(row)>8 else False  # IsZeroBonus
        fees_paid  = _b(row[9]) if len(row)>9 else False  # FeesPaidNonEnrolled
        note       = _s(row[1]) if len(row)>1 else ""

        # Auto-detect flags from status text
        sl = status.lower()
        if "enrolled, then visa granted" in sl: is_co=True
        if "current - enrolled" in sl: is_cur=True
        if "cancelled" in sl and "enrolled, then visa granted" not in sl: is_zero=True
        if "refused" in sl and "refused then granted" not in sl and "institution" not in sl: is_zero=True

        # Build conditions text
        conditions_parts = []
        if is_co: conditions_parts.append("Carry-over: student enrolled last month, visa arrived this month. CO receives deferred 50% at the prior month's tier rate.")
        if is_cur: conditions_parts.append("Current-Enrolled: student enrolled but visa still pending. CO receives 50% advance now; remaining 50% paid when visa/closure arrives.")
        if is_zero: conditions_parts.append("Zero bonus: application cancelled or visa refused. No payment made unless service fee was already collected (check FeesPaidNonEnrolled).")
        if fees_paid: conditions_parts.append("Service fee collected even though not enrolled — special handling applies (see Out-system/fee-paid column in base rates).")
        if not conditions_parts: conditions_parts.append(f"Standard case. Coun={int(coun_pct*100)}%, CO Direct={int(co_dir_pct*100)}%, CO Sub={int(co_sub_pct*100)}%.")

        db.add(StatusRule(
            status_value=status,
            is_eligible=not is_zero,
            counts_as_enrolled=counts_enrol,
            coun_pct=coun_pct,
            co_direct_pct=co_dir_pct,
            co_sub_pct=co_sub_pct,
            is_carry_over=is_co,
            is_current_enrolled=is_cur,
            is_zero_bonus=is_zero,
            fees_paid_non_enrolled=fees_paid,
            requires_visa="visa" in sl,
            requires_enrol=counts_enrol,
            dedup_rank=rank_map.get(sl, 0),
            start_date=EFFECTIVE_DATE,
            conditions=" | ".join(conditions_parts),
            triggers=f"CRM field 'Application Report Status' = '{status}'",
            note=note,
        ))
        count+=1

    for i, status in enumerate(status_order):
        db.add(ReferenceList(list_name="application_status", value=status, sort_order=i, is_active=True))

    db.flush()
    print(f"  05_STATUS_RULES:      {count} rules + split percentages + {len(status_order)} dropdown values")


# =============================================================================
# 09_SERVICE_FEE_RATES
# =============================================================================
def import_service_fee_rates(wb, db):
    ws = wb["09_SERVICE_FEE_RATES"]
    db.query(ServiceFeeRate).delete()
    db.query(ReferenceList).filter(ReferenceList.list_name=="service_fee_type").delete()
    db.flush()
    count=0; seen=set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        code=_s(row[0])
        if not code or code.startswith("(") or "rows" in code.lower() or code in seen: continue
        seen.add(code)
        cat=_s(row[5]).upper() if len(row)>5 else ""
        if not cat: continue
        db.add(ServiceFeeRate(fee_type=code, rate_pct=0.0,
                              flat_amount=_i(row[3]) if len(row)>3 else 0,
                              note=_s(row[7]) if len(row)>7 else "", is_active=True))
        db.add(ReferenceList(list_name="service_fee_type", value=code, sort_order=count, is_active=True))
        count+=1
    db.flush()
    print(f"  09_SERVICE_FEE_RATES: {count} records")


# =============================================================================
# 11_MASTER_AGENTS
# =============================================================================
def import_master_agents(wb, db):
    ws = wb["11_MASTER_AGENTS"]
    db.query(MasterAgent).delete(); db.flush()
    count=0; seen=set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        name=_s(row[1]); cls=_s(row[2])
        if not name or not cls or name in seen: continue
        seen.add(name)
        agent_type="MASTER_AGENT" if "master" in cls.lower() else "GROUP" if "group" in cls.lower() else "DIRECT"
        db.add(MasterAgent(agent_name=name, agent_type=agent_type, office="", is_active=True))
        count+=1
    db.flush()
    print(f"  11_MASTER_AGENTS:     {count} records")


# =============================================================================
# 13_SKIP_LABELS
# =============================================================================
def import_skip_labels(wb, db):
    ws = wb["13_SKIP_LABELS"]
    db.query(ReferenceList).filter(ReferenceList.list_name=="skip_labels").delete(); db.flush()
    count=0; seen=set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        label=_s(row[0])
        if not label or label.startswith("(") or label.startswith("13_") or label in seen: continue
        seen.add(label)
        db.add(ReferenceList(list_name="skip_labels", value=label, sort_order=count, is_active=True))
        count+=1
    db.flush()
    print(f"  13_SKIP_LABELS:       {count} records")


# =============================================================================
# 03_PRIORITY_INSTNS
# =============================================================================
def import_priority_institutions(wb, db):
    ws = wb["03_PRIORITY_INSTNS"]
    db.query(PriorityInstitution).delete(); db.flush()
    count=0; seen=set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        country=_s(row[0]); name=_s(row[1])
        if not name or not country or name in seen: continue
        if name.startswith("(") or name.startswith("03_"): continue
        seen.add(name)
        db.add(PriorityInstitution(country_code=country, institution_name=name,
                                   annual_target=_i(row[2]), bonus_pct=_f(row[3]) if row[3] else 0.0,
                                   direct_target=_i(row[4]), sub_target=_i(row[5]) if len(row)>5 else 0,
                                   is_active=True))
        count+=1
    db.flush()
    print(f"  03_PRIORITY_INSTNS:   {count} institutions")


# =============================================================================
# 08_YTD_TRACKER
# =============================================================================
def import_ytd_tracker(wb, db):
    ws = wb["08_YTD_TRACKER"]
    db.query(YtdTracker).delete(); db.flush()
    import datetime as dt
    current_year = dt.datetime.now().year
    count=0
    for row in ws.iter_rows(min_row=3, values_only=True):
        inst_name=_s(row[0])
        if not inst_name or inst_name.startswith("("): continue
        for m in range(1,13):
            val=_i(row[m]) if len(row)>m and row[m] is not None else 0
            if val>0:
                db.add(YtdTracker(institution_name=inst_name, year=current_year, month=m, enrolment_count=val))
                count+=1
    db.flush()
    print(f"  08_YTD_TRACKER:       {count} monthly enrolment records")


# =============================================================================
# 06_CLIENT_WEIGHTS
# =============================================================================
def import_client_weights(wb, db):
    ws = wb["06_CLIENT_WEIGHTS"]
    db.query(ClientWeight).delete(); db.flush()
    count=0; seen=set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        display=_s(row[0]); canonical=_s(row[1])
        if not canonical or canonical in seen or "ENGINE" in canonical: continue
        seen.add(canonical)
        db.add(ClientWeight(canonical_code=canonical, display_name=display,
                            weight_direct=_f(row[2]) if row[2] else 0.0,
                            weight_referred=_f(row[3]) if row[3] else 0.0,
                            weight_master=_f(row[4]) if row[4] else 0.0,
                            weight_outsys=_f(row[5]) if len(row)>5 and row[5] else 0.0,
                            weight_outsys_usa=_f(row[6]) if len(row)>6 and row[6] else 0.0,
                            note=_s(row[7]) if len(row)>7 else "", is_active=True))
        count+=1
    db.flush()
    print(f"  06_CLIENT_WEIGHTS:    {count} records")


# =============================================================================
# 07_CONTRACT_BONUS
# =============================================================================
def import_contract_bonuses(wb, db):
    ws = wb["07_CONTRACT_BONUS"]
    db.query(ContractBonus).delete(); db.flush()
    count=0; seen=set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        name=_s(row[0])
        if not name or name in seen or name.startswith("(") or name.startswith("CONTRACT"): continue
        seen.add(name)
        db.add(ContractBonus(package_name=name, service_fee_vnd=_s(row[1]),
                             coun_bonus=_s(row[2]), co_bonus=_s(row[3]),
                             timing=_s(row[4]) if len(row)>4 else "",
                             note=_s(row[5]) if len(row)>5 else "", is_active=True))
        count+=1
    db.flush()
    print(f"  07_CONTRACT_BONUS:    {count} records")


# =============================================================================
# 09_ADVANCE_TRACKER
# =============================================================================
def import_advance_tracker(wb, db):
    ws = wb["09_ADVANCE_TRACKER"]
    db.query(AdvancePayment).delete(); db.flush()
    count=0
    for row in ws.iter_rows(min_row=2, values_only=True):
        contract_id=_s(row[0])
        if not contract_id or not contract_id.startswith("SLC"): continue
        period_dt=_dt(row[3])
        db.add(AdvancePayment(
            contract_id=contract_id,
            student_name=_s(row[1]),
            staff_name=_s(row[2]),
            period_month=period_dt.month if period_dt else None,
            period_year=period_dt.year if period_dt else None,
            advance_paid=_i(row[4]),
            status_at_payment=_s(row[5]),
            full_bonus_at_tier=_i(row[6]),
            payment_type="Advance",
            is_settled=False,
            recorded_at=_dt(row[7]) or datetime.utcnow()
        ))
        count+=1
    db.flush()
    print(f"  09_ADVANCE_TRACKER:   {count} advance payment records")


# =============================================================================
# 02_BASE_BONUS_RATES — NEW: loads all rates into ref_base_rates
# =============================================================================
def import_base_rates(wb, db):
    ws = wb["02_BASE_BONUS_RATES"]
    db.query(BaseRate).delete(); db.flush()
    count=0

    def add_rate(scheme, tier, role, amount, description, conditions=""):
        if amount <= 0: return
        db.add(BaseRate(scheme=scheme, tier=tier, role=role, amount=amount,
                        start_date=EFFECTIVE_DATE, description=description,
                        conditions=conditions, is_active=True))
        nonlocal count; count+=1

    # Row 5: Section A — HCM Direct standard tiers
    # Cols: B=OutSysCoun, C=OutSysCO, D=VisaOnlyCO, E=UnderCoun, F=UnderCO,
    #       G=MeetHighCoun, H=MeetHighCO, I=MeetLowCoun, J=MeetLowCO, K=OverCoun, L=OverCO
    r=ws[5]
    add_rate("HCM_DIRECT","OUT_SYS","COUN", _i(r[1].value), "Out-system or fees paid but visa refused — Counsellor",
             "Applied when student is out-of-system and file closed (visa refused, cancelled) but service fee was collected.")
    add_rate("HCM_DIRECT","OUT_SYS","CO",   _i(r[2].value), "Out-system or fees paid but visa refused — CO",
             "Applied when student is out-of-system and file closed (visa refused, cancelled) but service fee was collected.")
    add_rate("HCM_DIRECT","VISA_ONLY","CO", _i(r[3].value), "Visa-only case — CO only",
             "Applied when application status = 'Closed - Visa granted (visa only)'. No enrolment; CO only receives bonus.")
    add_rate("HCM_DIRECT","UNDER","COUN",   _i(r[4].value), "Under target — Counsellor",
             "Applied when enrolled_count < monthly_target for the run month.")
    add_rate("HCM_DIRECT","UNDER","CO",     _i(r[5].value), "Under target — CO",
             "Applied when enrolled_count < monthly_target for the run month.")
    add_rate("HCM_DIRECT","MEET_LOW","COUN", _i(r[6].value), "Meet target, total bonus >= 5M — Counsellor",
             "Applied when enrolled_count = monthly_target AND sum(enrolled_bonuses) >= INCENTIVE_THRESHOLD (5,000,000 VND).")
    add_rate("HCM_DIRECT","MEET_LOW","CO",   _i(r[7].value), "Meet target, total bonus >= 5M — CO",
             "Applied when enrolled_count = monthly_target AND sum(enrolled_bonuses) >= INCENTIVE_THRESHOLD (5,000,000 VND).")
    add_rate("HCM_DIRECT","MEET_HIGH","COUN",_i(r[8].value), "Meet target, total bonus < 5M — Counsellor",
             "Applied when enrolled_count = monthly_target AND sum(enrolled_bonuses) < INCENTIVE_THRESHOLD (5,000,000 VND).")
    add_rate("HCM_DIRECT","MEET_HIGH","CO",  _i(r[9].value), "Meet target, total bonus < 5M — CO",
             "Applied when enrolled_count = monthly_target AND sum(enrolled_bonuses) < INCENTIVE_THRESHOLD (5,000,000 VND).")
    add_rate("HCM_DIRECT","OVER","COUN",    _i(r[10].value), "Over target — Counsellor",
             "Applied when enrolled_count > monthly_target. Rate applies to ALL cases in the month.")
    add_rate("HCM_DIRECT","OVER","CO",      _i(r[11].value), "Over target — CO",
             "Applied when enrolled_count > monthly_target. Rate applies to ALL cases in the month.")

    # Row 37: Section C — HN/DN Direct standard tiers (same column layout as Section A)
    r=ws[37]
    for scheme in ("HN_DIRECT",):
        add_rate(scheme,"OUT_SYS","COUN", _i(r[1].value),  "Out-system / fees paid — Counsellor (HN/DN)")
        add_rate(scheme,"OUT_SYS","CO",   _i(r[2].value),  "Out-system / fees paid — CO (HN/DN)")
        add_rate(scheme,"VISA_ONLY","CO", _i(r[3].value),  "Visa-only — CO (HN/DN)")
        add_rate(scheme,"UNDER","COUN",   _i(r[4].value),  "Under target — Counsellor (HN/DN)")
        add_rate(scheme,"UNDER","CO",     _i(r[5].value),  "Under target — CO (HN/DN)")
        add_rate(scheme,"MEET_LOW","COUN",_i(r[6].value),  "Meet target >= 5M — Counsellor (HN/DN)")
        add_rate(scheme,"MEET_LOW","CO",  _i(r[7].value),  "Meet target >= 5M — CO (HN/DN)")
        add_rate(scheme,"MEET_HIGH","COUN",_i(r[8].value), "Meet target < 5M — Counsellor (HN/DN)")
        add_rate(scheme,"MEET_HIGH","CO", _i(r[9].value),  "Meet target < 5M — CO (HN/DN)")
        add_rate(scheme,"OVER","COUN",    _i(r[10].value), "Over target — Counsellor (HN/DN)")
        add_rate(scheme,"OVER","CO",      _i(r[11].value), "Over target — CO (HN/DN)")

    # Section B: CO Sub-agent rates
    # Row 22: AUS/NZ/SG/US/CAN/UK — cols B=Partner/Fixed, C=EnrolOnly Under, D=Meet, E=Over, F=Full Under, G=Full Meet, H=Full Over
    r=ws[22]
    add_rate("CO_SUB","PARTNER","CO",      _i(r[1].value), "Enrol only thru partner (out-sys) — CO Sub",
             "Applied when institution name contains * or **. Sub-agent handles visa; CO does enrolment admin only.")
    add_rate("CO_SUB","UNDER","CO",        _i(r[2].value), "Enrolment only in-system — Under target — CO Sub",
             "Sub-agent provided the visa; CO handles enrolment only. Rate for under-target months.")
    add_rate("CO_SUB","MEET_LOW","CO",     _i(r[3].value), "Enrolment only in-system — Meet target — CO Sub",
             "Sub-agent provided the visa; CO handles enrolment only. Rate for meet-target months.")
    add_rate("CO_SUB","OVER","CO",         _i(r[4].value), "Enrolment only in-system — Over target — CO Sub",
             "Sub-agent provided the visa; CO handles enrolment only. Rate for over-target months.")
    add_rate("CO_SUB","UNDER_FULL","CO",   _i(r[5].value), "Full service in-system — Under target — CO Sub",
             "CO handles both enrolment and visa. Rate for under-target months.")
    add_rate("CO_SUB","MEET_LOW_FULL","CO",_i(r[6].value), "Full service in-system — Meet target — CO Sub",
             "CO handles both enrolment and visa. Rate for meet-target months.")
    add_rate("CO_SUB","OVER_FULL","CO",    _i(r[7].value), "Full service in-system — Over target — CO Sub",
             "CO handles both enrolment and visa. Rate for over-target months.")

    db.flush()
    print(f"  02_BASE_BONUS_RATES:  {count} rate records (HCM Direct + HN Direct + CO Sub)")


# =============================================================================
# SPECIAL RATES — NEW: Vietnam domestic, summer, guardian, visa-type fixed rates
# =============================================================================
def import_special_rates(wb, db):
    ws = wb["02_BASE_BONUS_RATES"]
    db.query(SpecialRate).delete(); db.flush()
    count = 0

    def add_sr(code, name, scheme, country_code, inst_pattern, ct_code, role, amount, conditions, description):
        if amount <= 0: return
        db.add(SpecialRate(rate_code=code, rate_name=name, scheme=scheme,
                           country_code=country_code, institution_pattern=inst_pattern,
                           client_type_code=ct_code, role=role, amount=amount,
                           start_date=EFFECTIVE_DATE, conditions=conditions,
                           description=description, is_active=True))
        nonlocal count; count+=1

    # --- Section A2 HCM (rows 7-18) ---
    # Row 7: THAI/PHIL/ML HCM Coun=1M, CO=500k
    r7=ws[7];  add_sr("THAI_PHIL_ML_HCM_COUN","Thai/Phil/ML flat — Counsellor HCM","HCM_DIRECT","THAI_PHIL_ML",None,None,"COUN",_i(r7[4].value),"country_code IN (TH, PH, MY). Flat rate, no target count.","Thailand, Philippines, Malaysia. Does NOT count toward monthly target.")
    r7=ws[7];  add_sr("THAI_PHIL_ML_HCM_CO","Thai/Phil/ML flat — CO HCM","HCM_DIRECT","THAI_PHIL_ML",None,None,"CO",_i(r7[5].value),"country_code IN (TH, PH, MY). Flat rate, no target count.","Thailand, Philippines, Malaysia. Does NOT count toward monthly target.")
    # Row 8: RMIT VN / BUV VN HCM
    r8=ws[8];  add_sr("RMIT_VN_HCM_COUN","RMIT VN / BUV VN undergrad/postgrad — HCM","HCM_DIRECT","VN","rmit|buv|british university",None,"COUN",_i(r8[4].value),"country = Vietnam AND institution contains 'RMIT' or 'BUV'. Fixed rate.","Vietnam domestic degree programs at RMIT Vietnam or British University Vietnam.")
    # Row 9: Other VN HCM
    r9=ws[9];  add_sr("OTHER_VN_HCM_COUN","Other Vietnam domestic programs — HCM","HCM_DIRECT","VN",None,None,"COUN",_i(r9[4].value),"country = Vietnam AND institution is NOT RMIT/BUV. Fixed rate.","All other Vietnam domestic programs including RMIT English, BUV English.")
    # Row 10: Summer HCM
    r10=ws[10]; add_sr("SUMMER_HCM_COUN","Summer study — HCM","HCM_DIRECT",None,None,"SUMMER_STUDY","COUN",_i(r10[4].value),"client_type_code = SUMMER_STUDY. Fixed rate, no target count.","Du học hè. Does NOT count toward monthly target.")
    # Row 11: Visa 485 HCM
    r11=ws[11]
    add_sr("VISA_485_HCM_COUN","Visa 485 — Counsellor HCM","HCM_DIRECT",None,None,"VISA_485","COUN",_i(r11[4].value),"service_fee_type or client_type = VISA_485","Graduate visa 485. Post-study work visa.")
    add_sr("VISA_485_HCM_CO","Visa 485 — CO HCM","HCM_DIRECT",None,None,"VISA_485","CO",_i(r11[5].value),"service_fee_type or client_type = VISA_485","Graduate visa 485.")
    # Row 12-13: Guardian granted/refused HCM
    r12=ws[12]; r13=ws[13]
    add_sr("GUARDIAN_VN_G_HCM_COUN","Guardian visa granted — Counsellor HCM","HCM_DIRECT",None,None,"GUARDIAN_VISA","COUN",_i(r12[4].value),"client_type_code = GUARDIAN_VISA AND status = granted","Guardian (Giám hộ) from Vietnam — granted outcome.")
    add_sr("GUARDIAN_VN_G_HCM_CO","Guardian visa granted — CO HCM","HCM_DIRECT",None,None,"GUARDIAN_VISA","CO",_i(r12[5].value),"client_type_code = GUARDIAN_VISA AND status = granted","Guardian from Vietnam — granted.")
    add_sr("GUARDIAN_VN_R_HCM_COUN","Guardian visa refused — Counsellor HCM","HCM_DIRECT",None,None,"GUARDIAN_VISA","COUN",_i(r13[4].value),"client_type_code = GUARDIAN_VISA AND status = refused","Guardian from Vietnam — refused.")
    add_sr("GUARDIAN_VN_R_HCM_CO","Guardian visa refused — CO HCM","HCM_DIRECT",None,None,"GUARDIAN_VISA","CO",_i(r13[5].value),"client_type_code = GUARDIAN_VISA AND status = refused","Guardian from Vietnam — refused.")
    # Row 14-15: Dependant
    r14=ws[14]; r15=ws[15]
    add_sr("DEPENDANT_G_HCM_COUN","Dependant granted — Counsellor HCM","HCM_DIRECT",None,None,"DEPENDANT_VISA","COUN",_i(r14[4].value),"client_type_code = DEPENDANT_VISA AND status = granted","Người phụ thuộc — visa granted.")
    add_sr("DEPENDANT_G_HCM_CO","Dependant granted — CO HCM","HCM_DIRECT",None,None,"DEPENDANT_VISA","CO",_i(r14[5].value),"client_type_code = DEPENDANT_VISA AND status = granted","Dependant — granted.")
    add_sr("DEPENDANT_R_HCM_COUN","Dependant refused — Counsellor HCM","HCM_DIRECT",None,None,"DEPENDANT_VISA","COUN",_i(r15[4].value),"client_type_code = DEPENDANT_VISA AND status = refused","Dependant — refused.")
    add_sr("DEPENDANT_R_HCM_CO","Dependant refused — CO HCM","HCM_DIRECT",None,None,"DEPENDANT_VISA","CO",_i(r15[5].value),"client_type_code = DEPENDANT_VISA AND status = refused","Dependant — refused.")
    # Row 16-18: Renewal / Guardian from AUS
    r16=ws[16]; r17=ws[17]; r18=ws[18]
    add_sr("STUDENT_VISA_RENEWAL_CO","Student visa renewal/extension AUS/NZ — CO","HCM_DIRECT",None,None,"STUDENT_VISA_RENEWAL","CO",_i(r16[5].value),"service_fee_type = STUDENT_VISA_RENEWAL","Student visa renewal or extension (AUS/NZ).")
    add_sr("VISITOR_EXCHANGE_CO","Visitor/Exchange/Business/Study permit renewal — CO","HCM_DIRECT",None,None,"VISITOR_EXCHANGE","CO",_i(r17[5].value),"service_fee_type = VISITOR_EXCHANGE or similar","Visitor, exchange, business, study permit renewal.")
    add_sr("GUARDIAN_AU_CO","Guardian from AUS (Aug 2022+) — CO","HCM_DIRECT",None,None,"GUARDIAN_VISA","CO",_i(r18[5].value),"client_type_code = GUARDIAN_VISA AND country_code = AU AND contract_date >= 2022-08-01","Guardian arrangement where guardian is based in Australia. 250k split 50/50.")

    # --- Section B CO Sub special rates (rows 23-32) ---
    r23=ws[23]
    add_sr("THAI_PHIL_ML_SUB","Thai/Phil/ML — CO Sub","CO_SUB","THAI_PHIL_ML",None,None,"CO",_i(r23[1].value),"country IN (Thailand, Philippines, Malaysia). CO Sub scheme. Fixed rate.","Flat rate for CO Sub scheme in these countries.")
    add_sr("THAI_PHIL_ML_SUB_UNDER","Thai/Phil/ML Under — CO Sub","CO_SUB","THAI_PHIL_ML",None,None,"CO",_i(r23[2].value),"country IN (Thailand, Philippines, Malaysia). CO Sub. Under target.","Under-target rate for CO Sub in Thailand/Philippines/Malaysia.")
    r24=ws[24]; add_sr("RMIT_VN_SUB","RMIT VN / BUV VN — CO Sub","CO_SUB","VN","rmit|buv",None,"CO",_i(r24[2].value),"country = Vietnam AND institution = RMIT/BUV. CO Sub scheme.","Vietnam domestic degree programs, CO Sub scheme.")
    r25=ws[25]; add_sr("OTHER_VN_SUB","Other VN programs — CO Sub","CO_SUB","VN",None,None,"CO",_i(r25[2].value),"country = Vietnam, not RMIT/BUV. CO Sub scheme.","All other Vietnam domestic programs, CO Sub scheme.")
    r26=ws[26]; add_sr("SUMMER_SUB","Summer study — CO Sub","CO_SUB",None,None,"SUMMER_STUDY","CO",_i(r26[2].value),"client_type_code = SUMMER_STUDY. CO Sub scheme.","Summer study, CO Sub scheme.")
    r27=ws[27]; add_sr("GUARDIAN_VN_G_SUB","Guardian granted — CO Sub","CO_SUB",None,None,"GUARDIAN_VISA","CO",_i(r27[1].value),"client_type_code = GUARDIAN_VISA AND granted. CO Sub.","Guardian visa granted, CO Sub scheme.")
    r28=ws[28]; add_sr("GUARDIAN_VN_R_SUB","Guardian refused — CO Sub","CO_SUB",None,None,"GUARDIAN_VISA","CO",_i(r28[1].value),"client_type_code = GUARDIAN_VISA AND refused. CO Sub.","Guardian visa refused, CO Sub scheme.")
    r29=ws[29]; add_sr("DEPENDANT_G_SUB","Dependant granted — CO Sub","CO_SUB",None,None,"DEPENDANT_VISA","CO",_i(r29[1].value),"client_type_code = DEPENDANT_VISA AND granted. CO Sub.","Dependant granted, CO Sub scheme.")
    r30=ws[30]; add_sr("DEPENDANT_R_SUB","Dependant refused — CO Sub","CO_SUB",None,None,"DEPENDANT_VISA","CO",_i(r30[1].value),"client_type_code = DEPENDANT_VISA AND refused. CO Sub.","Dependant refused, CO Sub scheme.")
    r31=ws[31]; add_sr("GUARDIAN_AU_SUB","Guardian from AUS (Aug 2022+) — CO Sub","CO_SUB",None,None,"GUARDIAN_VISA","CO",_i(r31[5].value),"client_type_code = GUARDIAN_VISA AND country = AU AND contract_date >= 2022-08-01. CO Sub.","Guardian from AUS, CO Sub scheme.")
    r32=ws[32]; add_sr("VISITOR_EXCHANGE_SUB","Visitor/Exchange/Other admin — CO Sub","CO_SUB",None,None,"VISITOR_EXCHANGE","CO",_i(r32[1].value),"service_fee_type = VISITOR_EXCHANGE or similar. CO Sub.","Visitor, exchange, other admin — CO Sub.")

    # --- Section C2 HN/DN special rates (rows 39-50) ---
    r39=ws[39]
    add_sr("THAI_PHIL_ML_HN_COUN","Thai/Phil/ML — Counsellor HN/DN","HN_DIRECT","THAI_PHIL_ML",None,None,"COUN",_i(r39[4].value),"country IN (TH, PH, MY). HN/DN office. Flat rate.","Thailand/Philippines/Malaysia, HN/DN office.")
    add_sr("THAI_PHIL_ML_HN_CO","Thai/Phil/ML — CO HN/DN","HN_DIRECT","THAI_PHIL_ML",None,None,"CO",_i(r39[5].value),"country IN (TH, PH, MY). HN/DN office. Flat rate.","Thailand/Philippines/Malaysia, HN/DN office.")

    db.flush()
    print(f"  02_SPECIAL_RATES:     {count} special rate records (HCM + HN + CO Sub)")


# =============================================================================
# INCENTIVE TIERS — NEW: seeds the 5M Meet threshold
# =============================================================================
def import_incentive_tiers(db):
    db.query(IncentiveTier).delete(); db.flush()
    db.add(IncentiveTier(
        type="MEET_THRESHOLD",
        name="Meet target — incentive threshold split (5M VND)",
        threshold_amount=5_000_000,
        service_types=json.dumps(["ALL"]),
        package_types=json.dumps(["ALL"]),
        start_date=EFFECTIVE_DATE,
        description=(
            "When a counsellor meets their monthly target, the total enrolled bonus "
            "for the month is compared to this threshold. "
            "If total >= threshold: MEET_LOW rate applies (lower per-case rate — high-value month). "
            "If total < threshold: MEET_HIGH rate applies (higher per-case rate — compensates for lower-value cases). "
            "Source: 02_BASE_BONUS_RATES columns G-H vs I-J."
        ),
        is_active=True,
    ))
    db.flush()
    print(f"  REF_INCENTIVE_TIERS:  1 record (Meet threshold 5,000,000 VND)")


# =============================================================================
# COUNTRY RATES — NEW: flat-rate countries per country
# =============================================================================
def import_country_rates(wb, db):
    ws = wb["02_BASE_BONUS_RATES"]
    db.query(CountryRate).delete(); db.flush()
    count=0

    flat_countries = [
        # (country_name, country_code, scheme, co_amount, coun_amount, description)
        ("Thailand",    "TH", "HCM_DIRECT", 500_000,  1_000_000, "Thailand — HCM Direct. Flat rate, does not count toward target."),
        ("Philippines", "PH", "HCM_DIRECT", 500_000,  1_000_000, "Philippines — HCM Direct. Flat rate, does not count toward target."),
        ("Malaysia",    "MY", "HCM_DIRECT", 500_000,  1_000_000, "Malaysia — HCM Direct. Flat rate, does not count toward target."),
        ("Thailand",    "TH", "HN_DIRECT",  400_000,  800_000,   "Thailand — HN/DN office. Flat rate, does not count toward target."),
        ("Philippines", "PH", "HN_DIRECT",  400_000,  800_000,   "Philippines — HN/DN office. Flat rate, does not count toward target."),
        ("Malaysia",    "MY", "HN_DIRECT",  400_000,  800_000,   "Malaysia — HN/DN office. Flat rate, does not count toward target."),
        ("Thailand",    "TH", "CO_SUB",     300_000,  0,         "Thailand — CO Sub scheme. Fixed rate."),
        ("Philippines", "PH", "CO_SUB",     300_000,  0,         "Philippines — CO Sub scheme. Fixed rate."),
        ("Malaysia",    "MY", "CO_SUB",     300_000,  0,         "Malaysia — CO Sub scheme. Fixed rate."),
    ]

    for country_name, country_code, scheme, co_amount, coun_amount, desc in flat_countries:
        db.add(CountryRate(
            country_name=country_name,
            country_code=country_code,
            scheme=scheme,
            rate_type="FLAT",
            co_amount=co_amount,
            coun_amount=coun_amount,
            counts_toward_target=False,
            start_date=EFFECTIVE_DATE,
            conditions=f"Applied when case country matches '{country_name}' (or '{country_code}') AND scheme = {scheme}. Bypasses tier calculation entirely.",
            description=desc,
            is_active=True,
        ))
        count+=1

    db.flush()
    print(f"  REF_COUNTRY_RATES:    {count} flat-rate country records")


# =============================================================================
# PARTNER INSTITUTIONS — NEW: seeds * and ** rate defaults
# =============================================================================
def import_partner_instns(db):
    db.query(PartnerInstitution).delete(); db.flush()

    # Single star (*) — out-of-system, sub-agent referral, enrolment only
    db.add(PartnerInstitution(
        partner_level="SINGLE",
        flag_pattern="*",
        rate_name="Out-system partner — enrolment only (*)",
        co_amount=400_000,
        coun_amount=0,
        start_date=EFFECTIVE_DATE,
        conditions=(
            "Institution name in CRM contains a single asterisk (*). "
            "The sub-agent referred the student and handles the visa; "
            "the CO only handles the enrolment side. "
            "Fixed rate bypasses tier calculation. "
            "Note: ** institutions also contain * and will match this rule unless "
            "the ** rule is checked first (which it is)."
        ),
        description="Standard out-of-system partner. CO receives 400,000 VND fixed. Source: 02_BASE_BONUS_RATES Section B row 22 col B.",
        is_active=True,
    ))

    # Double star (**) — premium partner (not yet separately coded in VBA, provisioned for future use)
    db.add(PartnerInstitution(
        partner_level="DOUBLE",
        flag_pattern="**",
        rate_name="Premium out-system partner — enrolment only (**)",
        co_amount=400_000,   # same as * until a different rate is confirmed
        coun_amount=0,
        start_date=EFFECTIVE_DATE,
        conditions=(
            "Institution name in CRM contains double asterisk (**). "
            "Indicates a premium-tier partner. "
            "Currently receives the same rate as single-star partners. "
            "Update co_amount here when a differentiated rate is confirmed."
        ),
        description="Premium out-of-system partner. Co_amount matches single-star for now. Update when policy differentiates ** rate.",
        is_active=True,
    ))

    db.flush()
    print(f"  REF_PARTNER_INSTNS:   2 records (* and ** partner levels)")


# =============================================================================
# ADVANCE RULES — NEW: seeds the default 50% advance rule
# =============================================================================
def import_advance_rules(db):
    db.query(AdvanceRule).delete(); db.flush()

    db.add(AdvanceRule(
        rule_name="Default 50% advance — Current-Enrolled",
        advance_pct=0.5,
        trigger_status="Current - Enrolled",
        service_type="ALL",
        country_code="ALL",
        institution_pattern=None,
        client_type_code="ALL",
        sort_order=100,
        start_date=EFFECTIVE_DATE,
        conditions=(
            "Applied when application status = 'Current - Enrolled' (student enrolled, visa still pending). "
            "CO receives 50% of the full tier bonus as an advance. "
            "Remaining 50% is paid when the file closes (visa granted or enrolment confirmed). "
            "If total advances exceed the final bonus (recovery scenario), "
            "net_payable = 0 and the overpayment is flagged for recovery."
        ),
        description="Default advance payment rule. Pays 50% upfront for current-enrolled cases. Source: 01_POLICY_SUMMARY Section III.",
        is_active=True,
    ))

    db.flush()
    print(f"  REF_ADVANCE_RULES:    1 record (default 50% advance)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_engine_config.py path/to/engine.xlsm")
        sys.exit(1)
    xlsm_path = sys.argv[1]
    if not os.path.exists(xlsm_path):
        print(f"File not found: {xlsm_path}")
        sys.exit(1)
    import_all(xlsm_path)
